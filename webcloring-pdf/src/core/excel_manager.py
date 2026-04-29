"""
Excel 파일 관리 클래스
자재 요청 데이터를 Excel 파일에 저장하고 관리합니다.
"""
import os
from pathlib import Path
from typing import Dict, List, Optional
from datetime import datetime, timedelta

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils.dataframe import dataframe_to_rows
except ImportError:
    print("openpyxl 패키지가 필요합니다. pip install openpyxl")
    raise

from config.settings import settings
from utils.logger import logger



class ExcelManager:
    """Excel 파일 관리 클래스 (실시간 모니터링용 개선)"""
    
    def __init__(self, file_path: Optional[str] = None):
        self.file_path = Path(file_path) if file_path else settings.excel_file_path
        self.workbook = None
        self.worksheet = None
        self.current_row = 1

        # 컬럼 정의
        self.columns = settings.get("excel.columns", [
            "순번", "자재코드", "품명", "요청수량(g단위)",
            "사유", "요청부서", "기안자", "문서번호", "처리일시"
        ])

        # 실시간 처리를 위한 추가 속성
        self.processed_documents = set()  # 처리된 문서 ID 추적
        self.last_save_time = None
        self.auto_save_interval = settings.auto_save_interval  # 설정 파일에서 로드

        self._initialize_workbook()
        self._load_processed_documents()
    
    def _initialize_workbook(self):
        """워크북 초기화"""
        try:
            # 기존 파일이 있으면 로드, 없으면 새로 생성
            if self.file_path.exists():
                self.workbook = openpyxl.load_workbook(self.file_path)
                self.worksheet = self.workbook.active
                
                # 마지막 행 찾기
                self.current_row = self.worksheet.max_row + 1
                
                # 빈 행이면 헤더부터 시작
                if self.current_row == 2 and not self.worksheet.cell(1, 1).value:
                    self.current_row = 1
                    
                logger.info(f"기존 Excel 파일 로드: {self.file_path}")
            else:
                self.workbook = openpyxl.Workbook()
                self.worksheet = self.workbook.active
                self.worksheet.title = "자재요청목록"
                self.current_row = 1
                
                logger.info(f"새 Excel 파일 생성: {self.file_path}")
            
            # 헤더가 없으면 생성
            if self.current_row == 1 or not self.worksheet.cell(1, 1).value:
                self._create_header()
                self.save()  # 헤더 생성 후 파일 저장

        except Exception as e:
            logger.error(f"Excel 워크북 초기화 실패: {e}")
            raise
    
    def _create_header(self):
        """헤더 행 생성"""
        try:
            # 헤더 텍스트 입력
            for col_idx, column_name in enumerate(self.columns, 1):
                cell = self.worksheet.cell(row=1, column=col_idx)
                cell.value = column_name
                
                # 헤더 스타일 적용
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = Border(
                    left=Side(style="thin"),
                    right=Side(style="thin"),
                    top=Side(style="thin"),
                    bottom=Side(style="thin")
                )
            
            # 컬럼 너비 자동 조정
            self._adjust_column_widths()
            
            self.current_row = 2
            logger.info("Excel 헤더 생성 완료")
            
        except Exception as e:
            logger.error(f"헤더 생성 실패: {e}")
    
    def _adjust_column_widths(self):
        """컬럼 너비 자동 조정"""
        column_widths = {
            "순번": 6,
            "자재코드": 15,
            "품명": 25,
            "요청수량(g단위)": 15,
            "사유": 20,
            "요청부서": 15,
            "기안자": 12,
            "문서번호": 20,
            "처리일시": 18,
            "PDF경로": 30
        }
        
        for col_idx, column_name in enumerate(self.columns, 1):
            width = column_widths.get(column_name, 12)
            column_letter = openpyxl.utils.get_column_letter(col_idx)
            self.worksheet.column_dimensions[column_letter].width = width
    
    def add_row(self, data: Dict[str, str]):
        """새 행 추가"""
        try:
            # 순번은 데이터에서 가져오거나 자동 생성
            sequence_number = data.get('순번', self.current_row - 1)
            
            # 데이터 매핑
            row_data = [
                sequence_number,  # 순번
                data.get('자재코드', ''),
                data.get('품명', ''),
                data.get('요청수량(g단위)', ''),
                data.get('사유', ''),
                data.get('요청부서', ''),
                data.get('기안자', ''),
                data.get('문서번호', ''),
                data.get('처리일시', datetime.now().strftime('%Y-%m-%d %H:%M:%S')),
                data.get('PDF경로', '')
            ]
            
            # 셀에 데이터 입력
            for col_idx, value in enumerate(row_data, 1):
                cell = self.worksheet.cell(row=self.current_row, column=col_idx)
                cell.value = value
                
                # 데이터 행 스타일 적용
                cell.alignment = Alignment(horizontal="left", vertical="center")
                cell.border = Border(
                    left=Side(style="thin"),
                    right=Side(style="thin"),
                    top=Side(style="thin"),
                    bottom=Side(style="thin")
                )
                
                # 순번 컬럼은 가운데 정렬
                if col_idx == 1:
                    cell.alignment = Alignment(horizontal="center", vertical="center")
            
            self.current_row += 1
            logger.debug(f"Excel 행 추가: {data.get('품명', 'Unknown')}")
            
        except Exception as e:
            logger.error(f"Excel 행 추가 실패: {e}")
    
    def add_multiple_rows(self, data_list: List[Dict[str, str]]):
        """여러 행 일괄 추가"""
        try:
            for data in data_list:
                self.add_row(data)
            
            logger.info(f"Excel 일괄 추가: {len(data_list)}행")
            
        except Exception as e:
            logger.error(f"Excel 일괄 추가 실패: {e}")
    
    def save(self, backup: bool = True):
        """Excel 파일 저장 (원자적 쓰기 적용)
        
        임시 파일로 저장 후 rename하여 원자적 교체를 보장합니다.
        저장 중 오류가 발생해도 기존 파일이 손상되지 않습니다.
        """
        import tempfile
        import shutil
        
        try:
            # 백업 생성
            if backup and self.file_path.exists():
                self._create_backup()
            
            # 디렉토리 생성 (필요한 경우)
            self.file_path.parent.mkdir(parents=True, exist_ok=True)
            
            # 1. 임시 파일로 저장
            temp_fd, temp_path = tempfile.mkstemp(
                suffix='.xlsx',
                dir=self.file_path.parent,
                prefix='.tmp_'
            )
            os.close(temp_fd)  # 파일 디스크립터 닫기
            
            # 워크북을 임시 파일에 저장
            self.workbook.save(temp_path)
            
            # 2. 원자적 교체 (rename)
            # Windows에서는 대상 파일이 있으면 먼저 삭제 필요
            if os.name == 'nt' and self.file_path.exists():
                os.replace(temp_path, self.file_path)
            else:
                shutil.move(temp_path, self.file_path)
            
            logger.file_saved(str(self.file_path), "Excel")
            logger.info(f"총 {self.current_row - 2}건의 데이터 저장 (원자적 쓰기)")
            
        except Exception as e:
            # 임시 파일 정리
            if 'temp_path' in locals() and os.path.exists(temp_path):
                try:
                    os.remove(temp_path)
                except:
                    pass
            logger.error(f"Excel 파일 저장 실패: {e}")
            raise

    
    def _create_backup(self):
        """백업 파일 생성"""
        try:
            backup_dir = settings.base_dir / settings.get("output.backup_directory", "backup")
            backup_dir.mkdir(exist_ok=True)
            
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            backup_name = f"{self.file_path.stem}_backup_{timestamp}{self.file_path.suffix}"
            backup_path = backup_dir / backup_name
            
            # 기존 파일을 백업으로 복사
            import shutil
            shutil.copy2(self.file_path, backup_path)
            
            logger.info(f"백업 생성: {backup_path}")
            
            # 오래된 백업 파일 정리 (최근 10개만 유지)
            self._cleanup_old_backups(backup_dir)
            
        except Exception as e:
            logger.warning(f"백업 생성 실패: {e}")
    
    def _cleanup_old_backups(self, backup_dir: Path, keep_count: int = 10):
        """오래된 백업 파일 정리"""
        try:
            backup_files = list(backup_dir.glob(f"{self.file_path.stem}_backup_*{self.file_path.suffix}"))
            
            if len(backup_files) > keep_count:
                # 수정 시간 기준 정렬 (최신 순)
                backup_files.sort(key=lambda x: x.stat().st_mtime, reverse=True)
                
                # 오래된 파일 삭제
                for old_file in backup_files[keep_count:]:
                    old_file.unlink()
                    logger.debug(f"오래된 백업 삭제: {old_file}")
                
        except Exception as e:
            logger.warning(f"백업 정리 실패: {e}")
    
    def get_existing_data(self) -> List[Dict[str, str]]:
        """기존 데이터 조회"""
        try:
            data_list = []
            
            if self.worksheet.max_row < 2:
                return data_list
            
            # 헤더 행 건너뛰고 데이터 읽기
            for row in self.worksheet.iter_rows(min_row=2, values_only=True):
                if not any(row):  # 빈 행 건너뛰기
                    continue
                
                data = {}
                for col_idx, value in enumerate(row):
                    if col_idx < len(self.columns):
                        column_name = self.columns[col_idx]
                        data[column_name] = str(value) if value is not None else ""
                
                data_list.append(data)
            
            return data_list
            
        except Exception as e:
            logger.error(f"기존 데이터 조회 실패: {e}")
            return []
    
    def find_duplicate(self, material_code: str, document_number: str) -> bool:
        """중복 데이터 확인"""
        try:
            existing_data = self.get_existing_data()
            
            for data in existing_data:
                if (data.get('자재코드') == material_code and 
                    data.get('문서번호') == document_number):
                    return True
            
            return False
            
        except Exception as e:
            logger.error(f"중복 확인 실패: {e}")
            return False
    
    def get_statistics(self) -> Dict[str, int]:
        """데이터 통계 정보 반환"""
        try:
            existing_data = self.get_existing_data()
            
            stats = {
                'total_records': len(existing_data),
                'unique_materials': len(set(data.get('자재코드', '') for data in existing_data if data.get('자재코드'))),
                'unique_documents': len(set(data.get('문서번호', '') for data in existing_data if data.get('문서번호'))),
                'unique_departments': len(set(data.get('요청부서', '') for data in existing_data if data.get('요청부서')))
            }
            
            return stats
            
        except Exception as e:
            logger.error(f"통계 정보 생성 실패: {e}")
            return {'total_records': 0, 'unique_materials': 0, 'unique_documents': 0, 'unique_departments': 0}
    
    def _load_processed_documents(self):
        """기존 파일에서 처리된 문서 ID 로드"""
        try:
            existing_data = self.get_existing_data()
            for data in existing_data:
                doc_number = data.get('문서번호', '')
                if doc_number:
                    self.processed_documents.add(doc_number)
            
            logger.info(f"기존 처리된 문서 로드: {len(self.processed_documents)}개")
            
        except Exception as e:
            logger.error(f"처리된 문서 로드 실패: {e}")
    
    def get_last_document_date(self) -> Optional[str]:
        """마지막 처리된 문서의 날짜 반환 (동적 필터링용)"""
        try:
            existing_data = self.get_existing_data()
            
            if not existing_data:
                return None
            
            # 처리일시 기준으로 가장 최근 데이터 찾기
            latest_date = None
            latest_datetime = None
            
            for data in existing_data:
                processed_time_str = data.get('처리일시', '')
                document_number = data.get('문서번호', '')
                
                if processed_time_str:
                    try:
                        # 처리일시에서 날짜 추출 (YYYY-MM-DD HH:MM:SS 형식)
                        processed_datetime = datetime.strptime(processed_time_str, '%Y-%m-%d %H:%M:%S')
                        
                        if latest_datetime is None or processed_datetime > latest_datetime:
                            latest_datetime = processed_datetime
                            latest_date = processed_datetime.strftime('%Y.%m.%d')
                    except ValueError:
                        pass
                
                # 처리일시가 없으면 문서번호에서 날짜 추출 시도
                if not latest_date and document_number:
                    try:
                        # 문서번호 형식: 20250724P001 등에서 날짜 추출
                        if len(document_number) >= 8 and document_number[:8].isdigit():
                            doc_date_str = document_number[:8]  # 20250724
                            doc_date = datetime.strptime(doc_date_str, '%Y%m%d')
                            
                            if latest_datetime is None or doc_date > latest_datetime:
                                latest_datetime = doc_date
                                latest_date = doc_date.strftime('%Y.%m.%d')
                    except ValueError:
                        pass
            
            if latest_date:
                logger.info(f"마지막 문서 날짜: {latest_date}")
                return latest_date
            else:
                logger.warning("마지막 문서 날짜를 찾을 수 없음")
                return None
                
        except Exception as e:
            logger.error(f"마지막 문서 날짜 조회 실패: {e}")
            return None
    
    def get_suggested_start_date(self, days_back: int = 0) -> str:
        """다음 검색에 사용할 권장 시작 날짜 반환"""
        try:
            last_date = self.get_last_document_date()
            
            if last_date:
                # 마지막 날짜를 기준으로 계산
                last_datetime = datetime.strptime(last_date, '%Y.%m.%d')

                # days_back 만큼 이전 날짜부터 시작 (놓칠 수 있는 문서 고려)
                start_datetime = last_datetime - timedelta(days=days_back)
                suggested_date = start_datetime.strftime('%Y.%m.%d')

                logger.info(f"권장 시작 날짜: {suggested_date} (마지막 문서: {last_date}, {days_back}일 전부터)")
                return suggested_date
            else:
                # 마지막 문서가 없으면 설정 날짜 사용
                if settings.search_start_date:
                    logger.info(f"Excel 데이터 없음 → 설정 날짜 사용: {settings.search_start_date}")
                    return settings.search_start_date
                else:
                    # 설정 날짜도 없으면 오류
                    error_msg = "스마트 필터링이 활성화되었지만 Excel 데이터도 없고 설정 날짜(SEARCH_START_DATE)도 없습니다."
                    logger.error(error_msg)
                    raise ValueError(error_msg)

        except ValueError:
            # ValueError는 그대로 전파 (설정 오류)
            raise
        except Exception as e:
            logger.error(f"권장 시작 날짜 계산 실패: {e}")
            # 기타 오류 시 설정 날짜로 폴백
            if settings.search_start_date:
                logger.warning(f"오류 발생, 설정 날짜로 폴백: {settings.search_start_date}")
                return settings.search_start_date
            else:
                raise ValueError("시작 날짜를 결정할 수 없습니다. SEARCH_START_DATE를 설정해주세요.")
    
    def add_document_data(self, document: Dict[str, any]):
        """문서 데이터 추가 (실시간 모니터링용)"""
        try:
            doc_id = document.get('id', '')
            
            # 중복 처리 방지
            if doc_id in self.processed_documents:
                logger.debug(f"이미 처리된 문서 무시: {doc_id}")
                return
            
            # 자재 정보가 있는 경우 각 자재별로 행 추가
            materials = document.get('materials', [])
            if materials:
                for material in materials:
                    material_data = {
                        '자재코드': material.get('material_code', ''),
                        '품명': material.get('material_name', ''),
                        '요청수량(g단위)': material.get('quantity', ''),
                        '사유': material.get('reason', ''),
                        '요청부서': document.get('department', ''),
                        '기안자': document.get('author', ''),
                        '문서번호': doc_id,
                        '처리일시': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                    }
                    self.add_row(material_data)
            else:
                # 자재 정보가 없는 경우 문서 기본 정보만 추가
                document_data = {
                    '자재코드': '',
                    '품명': document.get('title', ''),
                    '요청수량(g단위)': '',
                    '사유': '',
                    '요청부서': document.get('department', ''),
                    '기안자': document.get('author', ''),
                    '문서번호': doc_id
                }
                self.add_row(document_data)
            
            # 처리된 문서로 표시
            self.processed_documents.add(doc_id)
            
            # 자동 저장 확인
            self._check_auto_save()
            
            logger.info(f"문서 데이터 추가 완료: {doc_id}")
            
        except Exception as e:
            logger.error(f"문서 데이터 추가 실패: {e}")
    
    def _check_auto_save(self):
        """자동 저장 시간 확인 및 수행"""
        try:
            current_time = datetime.now()
            
            if (self.last_save_time is None or 
                (current_time - self.last_save_time).total_seconds() >= self.auto_save_interval):
                
                self.save(backup=False)  # 자동 저장시에는 백업 생성 안함
                self.last_save_time = current_time
                logger.debug("Excel 자동 저장 완료")
            
        except Exception as e:
            logger.error(f"자동 저장 실패: {e}")
    
    def is_document_processed(self, doc_id: str) -> bool:
        """문서 처리 여부 확인"""
        return doc_id in self.processed_documents
    
    def get_processing_stats(self) -> Dict[str, any]:
        """처리 통계 정보"""
        stats = self.get_statistics()
        stats.update({
            'processed_documents_count': len(self.processed_documents),
            'last_save_time': self.last_save_time.isoformat() if self.last_save_time else None,
            'auto_save_interval_minutes': self.auto_save_interval / 60
        })
        return stats
    
    def force_save(self):
        """강제 저장 (백업 포함)"""
        try:
            self.save(backup=True)
            self.last_save_time = datetime.now()
            logger.info("Excel 강제 저장 완료")
        except Exception as e:
            logger.error(f"강제 저장 실패: {e}")
    
    def reset_processed_documents(self):
        """처리된 문서 목록 초기화 (신중히 사용)"""
        self.processed_documents.clear()
        logger.warning("처리된 문서 목록이 초기화되었습니다")
    
    def save_material_data(self, table_rows, drafter, document_number, department):
        """자재 데이터 저장 (기존 portal_automation.py 호환용)"""
        try:
            logger.info(f"자재 데이터 저장 시작: 문서번호 {document_number}, {len(table_rows)}행")
            
            # 중복 문서 확인
            if document_number in self.processed_documents:
                logger.debug(f"이미 처리된 문서 무시: {document_number}")
                return
            
            # 각 자재 행을 Excel에 추가
            for row_data in table_rows:
                # row_data가 리스트인 경우 (셀 값들의 배열)
                if isinstance(row_data, list) and len(row_data) >= 4:
                    material_data = {
                        '순번': row_data[0] if len(row_data) > 0 else '',  # 원본 문서의 순번 사용
                        '자재코드': row_data[1] if len(row_data) > 1 else '',
                        '품명': row_data[2] if len(row_data) > 2 else '',
                        '요청수량(g단위)': row_data[3] if len(row_data) > 3 else '',
                        '사유': row_data[4] if len(row_data) > 4 else '',
                        '요청부서': department,
                        '기안자': drafter,
                        '문서번호': document_number,
                        '처리일시': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                    }
                    
                # row_data가 딕셔너리인 경우
                elif isinstance(row_data, dict):
                    material_data = {
                        '순번': row_data.get('순번', row_data.get('sequence', '')),
                        '자재코드': row_data.get('자재코드', row_data.get('material_code', '')),
                        '품명': row_data.get('품명', row_data.get('material_name', '')),
                        '요청수량(g단위)': row_data.get('요청수량(g단위)', row_data.get('quantity', '')),
                        '사유': row_data.get('사유', row_data.get('reason', '')),
                        '요청부서': department,
                        '기안자': drafter,
                        '문서번호': document_number,
                        '처리일시': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                    }
                
                else:
                    # 기본 처리
                    material_data = {
                        '자재코드': str(row_data),
                        '품명': '',
                        '요청수량(g단위)': '',
                        '사유': '',
                        '요청부서': department,
                        '기안자': drafter,
                        '문서번호': document_number,
                        '처리일시': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                    }
                
                # 빈 자재코드나 품명이 있는 행은 건너뛰기
                if not material_data['자재코드'] and not material_data['품명']:
                    continue
                    
                self.add_row(material_data)
            
            # 처리된 문서로 표시
            self.processed_documents.add(document_number)
            
            # 즉시 저장 (데이터 유실 방지)
            self.save(backup=False)
            
            logger.info(f"자재 데이터 저장 완료: 문서번호 {document_number}")
            
        except Exception as e:
            logger.error(f"자재 데이터 저장 실패: {e}")
            raise

    def close(self):
        """워크북 닫기"""
        try:
            # 마지막 저장
            if self.workbook:
                self.force_save()
                self.workbook.close()
                logger.debug("Excel 워크북 닫기 완료")
        except Exception as e:
            logger.warning(f"워크북 닫기 실패: {e}")

    # ==================== Google Sheets 통합 (후방 호환) ====================

    @property
    def wb(self):
        """GoogleSheetsManager 호환성을 위한 workbook 별칭"""
        return self.workbook

    def finalize_google_backup(self) -> bool:
        """Google Sheets 백업 (GoogleBackupManager에 위임)

        후방 호환성을 위해 유지됩니다.
        신규 코드에서는 GoogleBackupManager를 직접 사용하세요.
        """
        try:
            from services.google_backup_manager import GoogleBackupManager
            backup = GoogleBackupManager(self)
            return backup.finalize()
        except ImportError:
            logger.debug("GoogleBackupManager를 사용할 수 없습니다")
            return False
        except Exception as e:
            logger.error(f"Google Sheets 백업 중 오류: {e}")
            return False